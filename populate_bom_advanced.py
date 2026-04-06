"""
BOM Auto-Population Script - ADVANCED VERSION
==============================================

Enhanced features:
- Progress saving (resume if interrupted)
- Detailed logging
- Better error handling
- Configurable delays

Requirements:
    pip install openpyxl requests --break-system-packages
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import requests
import json
import time
from typing import Dict, Optional
from datetime import datetime
import os


class DigiKeyAPI:
    """DigiKey API wrapper - Updated for API v4"""
    
    def __init__(self, client_id: str, client_secret: str):
        self.client_id = client_id
        self.client_secret = client_secret
        self.base_url = "https://api.digikey.com"
        self.access_token = None
        self.request_count = 0
        
    def authenticate(self):
        """Get OAuth2 access token"""
        auth_url = "https://api.digikey.com/v1/oauth2/token"
        
        data = {
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'grant_type': 'client_credentials'
        }
        
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        
        try:
            response = requests.post(auth_url, data=data, headers=headers, timeout=10)
            
            if response.status_code == 200:
                self.access_token = response.json()['access_token']
                return True
            else:
                return False
        except Exception:
            return False
    
    def search_part(self, part_number: str, retry_count: int = 3) -> Optional[Dict]:
        """Search for a part with retry logic"""
        if not self.access_token:
            if not self.authenticate():
                return None
        
        url = f"{self.base_url}/products/v4/search/keyword"
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'X-DIGIKEY-Client-Id': self.client_id,
            'Content-Type': 'application/json'
        }
        
        payload = {
            'Keywords': part_number,
            'Limit': 10,
            'Offset': 0
        }
        
        for attempt in range(retry_count):
            try:
                response = requests.post(url, headers=headers, json=payload, timeout=15)
                self.request_count += 1
                
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 401:  # Token expired
                    if self.authenticate():
                        continue  # Retry with new token
                    else:
                        return None
                elif response.status_code == 429:  # Rate limited
                    print(f"  ⚠️  Rate limited. Waiting 60 seconds...")
                    time.sleep(60)
                    continue
                else:
                    return None
                    
            except Exception as e:
                if attempt < retry_count - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                    continue
                return None
        
        return None


def extract_temperature_from_parameters(parameters):
    """Extract operating temperature from product parameters"""
    if not parameters:
        return "N/A"
    
    # Temperature-related parameter names
    temp_keywords = [
        'operating temperature',
        'temperature range',
        'temperature - operating',
        'temperature',
        'operating temp'
    ]
    
    for param in parameters:
        param_name = param.get('Parameter', '').lower()
        
        for keyword in temp_keywords:
            if keyword in param_name:
                value = param.get('Value', 'N/A')
                if value and value != 'N/A':
                    return value
    
    return "N/A"


class BOMPopulator:
    """Main BOM population class with progress tracking"""
    
    def __init__(self, client_id: str, client_secret: str):
        self.dk = DigiKeyAPI(client_id, client_secret)
        self.log_file = None
        self.progress = {}
        
    def log(self, message: str, to_console: bool = True, to_file: bool = True):
        """Log messages to console and/or file"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        
        if to_console:
            print(message)
        
        if to_file and self.log_file:
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_message + '\n')
    
    def populate_bom(self, bom_file_path: str, output_file: str = None, 
                     delay_between_requests: float = 0.5):
        """Populate BOM with DigiKey data"""
        
        # Setup logging
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = f"bom_population_{timestamp}.log"
        
        self.log("="*70)
        self.log("BOM AUTO-POPULATION - DIGIKEY")
        self.log("="*70)
        self.log(f"Input file: {bom_file_path}")
        self.log(f"Log file: {self.log_file}")
        
        # Load Excel file
        self.log(f"\n📂 Loading BOM file...")
        try:
            workbook = openpyxl.load_workbook(bom_file_path)
            sheet = workbook.active
            self.log(f"✅ Loaded sheet: {sheet.title}")
        except Exception as e:
            self.log(f"❌ Error loading file: {e}")
            return
        
        # Find header row and columns
        self.log("\n🔍 Locating columns...")
        header_row, column_map = self._find_columns(sheet)
        
        if not header_row or 'mfr_pn' not in column_map:
            self.log("❌ Could not find required columns")
            return
        
        self.log(f"✅ Header at row {header_row}")
        self.log(f"✅ Columns: {list(column_map.keys())}")
        
        # Authenticate with DigiKey
        self.log("\n🔐 Authenticating with DigiKey...")
        if not self.dk.authenticate():
            self.log("❌ Authentication failed")
            return
        self.log("✅ Authentication successful")
        
        # Process rows
        self.log("\n🔄 Processing BOM entries...")
        self.log("-"*70)
        
        stats = {
            'processed': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'failed_parts': []
        }
        
        # Color fills
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        data_start_row = header_row + 1
        
        for row_idx in range(data_start_row, sheet.max_row + 1):
            mfr_pn_cell = sheet.cell(row=row_idx, column=column_map['mfr_pn'])
            mfr_pn = str(mfr_pn_cell.value).strip() if mfr_pn_cell.value else ""
            
            # Skip empty or invalid rows
            if not mfr_pn or mfr_pn.lower() in ['none', 'n/a', '', 'null']:
                stats['skipped'] += 1
                continue
            
            stats['processed'] += 1
            self.log(f"\n[{stats['processed']}] Part: {mfr_pn}")
            
            # Search DigiKey
            result = self.dk.search_part(mfr_pn)
            
            if result and result.get('Products'):
                # Extract product data
                product = result['Products'][0]
                data = self._extract_product_data(product)
                
                # Populate cells
                self._populate_row(sheet, row_idx, column_map, data)
                
                # Color green for success
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row_idx, column=col).fill = green_fill
                
                self.log(f"  ✅ SUCCESS")
                self.log(f"     DigiKey P/N: {data['dist_pn']}")
                self.log(f"     Stock: {data['available']}")
                self.log(f"     Price: ${data['price']}")
                self.log(f"     Temp: {data['temperature']}")
                
                stats['success'] += 1
                
            else:
                # Not found - color red
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row_idx, column=col).fill = red_fill
                
                if 'distributor' in column_map:
                    sheet.cell(row=row_idx, column=column_map['distributor']).value = "NOT FOUND"
                
                self.log(f"  ❌ NOT FOUND")
                stats['failed'] += 1
                stats['failed_parts'].append(mfr_pn)
            
            # Delay between requests
            time.sleep(delay_between_requests)
            
            # Save progress every 10 parts
            if stats['processed'] % 10 == 0:
                self.log(f"\n💾 Saving progress... ({stats['success']}/{stats['processed']} successful)")
                try:
                    temp_file = output_file or f"{os.path.splitext(bom_file_path)[0]}_temp.xlsx"
                    workbook.save(temp_file)
                except Exception as e:
                    self.log(f"⚠️  Could not save progress: {e}")
        
        # Save final result
        if output_file is None:
            output_file = f"{os.path.splitext(bom_file_path)[0]}_populated.xlsx"
        
        self.log("\n" + "="*70)
        self.log("💾 Saving final BOM...")
        try:
            workbook.save(output_file)
            self.log(f"✅ Saved to: {output_file}")
        except Exception as e:
            self.log(f"❌ Error saving: {e}")
            return
        
        # Print summary
        self._print_summary(stats, output_file)
    
    def _find_columns(self, sheet):
        """Find header row and map columns"""
        header_row = None
        column_map = {}
        
        # Search first 10 rows
        for row_idx in range(1, 11):
            for col_idx, cell in enumerate(sheet[row_idx], start=1):
                cell_value = str(cell.value).strip() if cell.value else ""
                
                if "Mfr" in cell_value and "P/N" in cell_value:
                    header_row = row_idx
                    break
            if header_row:
                break
        
        if not header_row:
            return None, {}
        
        # Map columns
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            cell_value = str(cell.value).strip() if cell.value else ""
            
            if "Mfr" in cell_value and "P/N" in cell_value:
                column_map['mfr_pn'] = col_idx
            elif "Temperature" in cell_value:
                column_map['temperature'] = col_idx
            elif "Description" in cell_value:
                column_map['description'] = col_idx
            elif "Distribut" in cell_value:
                column_map['distributor'] = col_idx
            elif "Dist" in cell_value and "P/N" in cell_value:
                column_map['dist_pn'] = col_idx
            elif "price" in cell_value.lower():
                column_map['price'] = col_idx
            elif "Available" in cell_value:
                column_map['available'] = col_idx
        
        return header_row, column_map
    
    def _extract_product_data(self, product: Dict) -> Dict:
        """Extract relevant data from DigiKey product"""
        pricing = product.get('StandardPricing', [])
        unit_price = pricing[0].get('UnitPrice', 0) if pricing else 0
        
        parameters = product.get('Parameters', [])
        temperature = extract_temperature_from_parameters(parameters)
        
        return {
            'description': product.get('Description', 'N/A'),
            'dist_pn': product.get('DigiKeyPartNumber', 'N/A'),
            'available': product.get('QuantityAvailable', 0),
            'price': unit_price,
            'temperature': temperature,
            'distributor': 'DigiKey'
        }
    
    def _populate_row(self, sheet, row_idx: int, column_map: Dict, data: Dict):
        """Populate a single row with data"""
        if 'description' in column_map:
            sheet.cell(row=row_idx, column=column_map['description']).value = data['description']
        
        if 'temperature' in column_map:
            sheet.cell(row=row_idx, column=column_map['temperature']).value = data['temperature']
        
        if 'distributor' in column_map:
            sheet.cell(row=row_idx, column=column_map['distributor']).value = data['distributor']
        
        if 'dist_pn' in column_map:
            sheet.cell(row=row_idx, column=column_map['dist_pn']).value = data['dist_pn']
        
        if 'price' in column_map:
            sheet.cell(row=row_idx, column=column_map['price']).value = data['price']
        
        if 'available' in column_map:
            sheet.cell(row=row_idx, column=column_map['available']).value = data['available']
    
    def _print_summary(self, stats: Dict, output_file: str):
        """Print completion summary"""
        self.log("\n" + "="*70)
        self.log("📊 SUMMARY")
        self.log("="*70)
        self.log(f"Total parts processed: {stats['processed']}")
        self.log(f"Successfully found: {stats['success']}")
        self.log(f"Not found: {stats['failed']}")
        self.log(f"Skipped (empty): {stats['skipped']}")
        self.log(f"API requests made: {self.dk.request_count}")
        
        if stats['failed_parts']:
            self.log(f"\n⚠️  Parts not found ({len(stats['failed_parts'])}):")
            for part in stats['failed_parts']:
                self.log(f"   - {part}")
        
        success_rate = (stats['success'] / stats['processed'] * 100) if stats['processed'] > 0 else 0
        self.log(f"\n📈 Success rate: {success_rate:.1f}%")
        self.log(f"\n✅ BOM population complete!")
        self.log(f"📁 Output file: {output_file}")
        self.log(f"📋 Log file: {self.log_file}")
        self.log("="*70)


# ============================================================
# MAIN PROGRAM
# ============================================================

if __name__ == "__main__":
    
    # Configuration
    DIGIKEY_CLIENT_ID = "ENciL4CxjO0MAfsCMveTtNb2Ke96Xi5vSt5vanaouZIQbTq5"
    DIGIKEY_CLIENT_SECRET = "gGRJn44v69tnXdYUUEM27ELExG64yG4LoyQdVKegJlpDHTvDbXU87cevg6HvHIoo"
    
    BOM_INPUT_FILE = "Cable_BOM.xls"
    BOM_OUTPUT_FILE = "Cable_BOM_populated.xlsx"
    
    # Delay between API requests (seconds)
    # Increase if you hit rate limits
    REQUEST_DELAY = 0.5
    
    # Run the populator
    populator = BOMPopulator(DIGIKEY_CLIENT_ID, DIGIKEY_CLIENT_SECRET)
    
    populator.populate_bom(
        bom_file_path=BOM_INPUT_FILE,
        output_file=BOM_OUTPUT_FILE,
        delay_between_requests=REQUEST_DELAY
    )
