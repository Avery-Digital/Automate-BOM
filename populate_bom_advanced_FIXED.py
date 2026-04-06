"""
BOM Auto-Population Script - ADVANCED VERSION
==============================================

Enhanced features:
- Progress saving (resume if interrupted)
- Detailed logging
- Better error handling
- Configurable delays

Requirements:
    pip install openpyxl requests --break-system-packages
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import requests
import json
import time
import urllib.parse
from typing import Dict, Optional
from datetime import datetime
import os


class DigiKeyAPI:
    """DigiKey API wrapper - Updated for API v4"""
    
    def __init__(self, client_id: str, client_secret: str):
        self.client_id = client_id
        self.client_secret = client_secret
        self.base_url = "https://api.digikey.com"
        self.access_token = None
        self.request_count = 0
        
    def authenticate(self):
        """Get OAuth2 access token"""
        auth_url = "https://api.digikey.com/v1/oauth2/token"
        
        data = {
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'grant_type': 'client_credentials'
        }
        
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        
        try:
            response = requests.post(auth_url, data=data, headers=headers, timeout=10)
            
            if response.status_code == 200:
                self.access_token = response.json()['access_token']
                return True
            else:
                return False
        except Exception:
            return False
    
    def search_part(self, part_number: str, retry_count: int = 3) -> Optional[Dict]:
        """Search for a part with retry logic"""
        if not self.access_token:
            if not self.authenticate():
                return None
        
        url = f"{self.base_url}/products/v4/search/keyword"
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'X-DIGIKEY-Client-Id': self.client_id,
            'Content-Type': 'application/json'
        }
        
        payload = {
            'Keywords': part_number,
            'Limit': 10,
            'Offset': 0
        }
        
        for attempt in range(retry_count):
            try:
                response = requests.post(url, headers=headers, json=payload, timeout=15)
                self.request_count += 1
                
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 401:  # Token expired
                    if self.authenticate():
                        continue  # Retry with new token
                    else:
                        return None
                elif response.status_code == 429:  # Rate limited
                    print(f"  ⚠️  Rate limited. Waiting 60 seconds...")
                    time.sleep(60)
                    continue
                else:
                    return None
                    
            except Exception as e:
                if attempt < retry_count - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                    continue
                return None
        
        return None
    
    def get_product_pricing(self, digikey_part_number: str, retry_count: int = 3) -> Optional[Dict]:
        """Get pricing for a specific DigiKey part number using the Product Pricing API"""
        if not self.access_token:
            if not self.authenticate():
                return None
        
        # URL encode the part number for the API call
        encoded_pn = urllib.parse.quote(digikey_part_number, safe='')
        
        url = f"{self.base_url}/products/v4/search/{encoded_pn}/productdetails"
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'X-DIGIKEY-Client-Id': self.client_id,
            'Content-Type': 'application/json'
        }
        
        for attempt in range(retry_count):
            try:
                response = requests.get(url, headers=headers, timeout=15)
                self.request_count += 1
                
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 401:  # Token expired
                    if self.authenticate():
                        continue  # Retry with new token
                    else:
                        return None
                elif response.status_code == 429:  # Rate limited
                    print(f"  ⚠️  Rate limited on pricing API. Waiting 60 seconds...")
                    time.sleep(60)
                    continue
                else:
                    return None
                    
            except Exception as e:
                if attempt < retry_count - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                    continue
                return None
        
        return None


def extract_temperature_from_parameters(parameters):
    """Extract operating temperature from product parameters"""
    if not parameters:
        return ""
    
    # Temperature-related parameter names
    temp_keywords = [
        'operating temperature',
        'temperature range',
        'temperature - operating',
        'temperature',
        'operating temp'
    ]
    
    for param in parameters:
        param_name = param.get('Parameter', '').lower()
        
        for keyword in temp_keywords:
            if keyword in param_name:
                value = param.get('Value', '')
                if value and value != 'N/A':
                    return value
    
    return ""


class BOMPopulator:
    """Main BOM population class with progress tracking"""
    
    def __init__(self, client_id: str, client_secret: str):
        self.dk = DigiKeyAPI(client_id, client_secret)
        self.log_file = None
        self.progress = {}
        
    def log(self, message: str, to_console: bool = True, to_file: bool = True):
        """Log messages to console and/or file"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        
        if to_console:
            print(message)
        
        if to_file and self.log_file:
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_message + '\n')
    
    def populate_bom(self, bom_file_path: str, output_file: str = None, 
                     delay_between_requests: float = 0.5):
        """Populate BOM with DigiKey data"""
        
        # Setup logging
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = f"bom_population_{timestamp}.log"
        
        self.log("="*70)
        self.log("BOM AUTO-POPULATION - DIGIKEY")
        self.log("="*70)
        self.log(f"Input file: {bom_file_path}")
        self.log(f"Log file: {self.log_file}")
        
        # Load Excel file
        self.log(f"\n📂 Loading BOM file...")
        try:
            workbook = openpyxl.load_workbook(bom_file_path)
            sheet = workbook.active
            self.log(f"✅ Loaded sheet: {sheet.title}")
        except Exception as e:
            self.log(f"❌ Error loading file: {e}")
            return
        
        # Find header row and columns
        self.log("\n🔍 Locating columns...")
        header_row, column_map = self._find_columns(sheet)
        
        if not header_row or 'mfr_pn' not in column_map:
            self.log("❌ Could not find required columns")
            return
        
        self.log(f"✅ Header at row {header_row}")
        self.log(f"✅ Columns: {list(column_map.keys())}")
        
        # Authenticate with DigiKey
        self.log("\n🔐 Authenticating with DigiKey...")
        if not self.dk.authenticate():
            self.log("❌ Authentication failed")
            return
        self.log("✅ Authentication successful")
        
        # Process rows
        self.log("\n🔄 Processing BOM entries...")
        self.log("-"*70)
        
        stats = {
            'processed': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'failed_parts': []
        }
        
        # Color fills
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        data_start_row = header_row + 1
        
        for row_idx in range(data_start_row, sheet.max_row + 1):
            mfr_pn_cell = sheet.cell(row=row_idx, column=column_map['mfr_pn'])
            mfr_pn = str(mfr_pn_cell.value).strip() if mfr_pn_cell.value else ""
            
            # Skip empty or invalid rows
            if not mfr_pn or mfr_pn.lower() in ['none', 'n/a', '', 'null']:
                stats['skipped'] += 1
                continue
            
            stats['processed'] += 1
            self.log(f"\n[{stats['processed']}] Part: {mfr_pn}")
            
            # Search DigiKey
            result = self.dk.search_part(mfr_pn)
            
            if result and result.get('Products'):
                # Find the best matching product - prioritize exact matches
                products = result['Products']
                product = None
                
                # First, try to find an exact match on ManufacturerProductNumber
                for p in products:
                    mpn = p.get('ManufacturerProductNumber', '')
                    if mpn.upper() == mfr_pn.upper():
                        product = p
                        self.log(f"  🎯 Found exact MPN match: {mpn}")
                        break
                
                # If no exact match, check ExactMatches array if available
                if not product and result.get('ExactMatches'):
                    for p in result['ExactMatches']:
                        mpn = p.get('ManufacturerProductNumber', '')
                        if mpn.upper() == mfr_pn.upper():
                            product = p
                            self.log(f"  🎯 Found exact match in ExactMatches: {mpn}")
                            break
                
                # If still no exact match, try partial match (MPN contains search term or vice versa)
                if not product:
                    for p in products:
                        mpn = p.get('ManufacturerProductNumber', '')
                        if mfr_pn.upper() in mpn.upper() or mpn.upper() in mfr_pn.upper():
                            product = p
                            self.log(f"  🔍 Found partial MPN match: {mpn}")
                            break
                
                # Fall back to first result only if nothing better found
                if not product:
                    product = products[0]
                    self.log(f"  ⚠️  No exact match, using first result: {product.get('ManufacturerProductNumber', 'N/A')}")
                
                # Get DigiKey part number to fetch pricing - it's in ProductVariations array
                digikey_pn = ''
                product_variations = product.get('ProductVariations', [])
                if product_variations and len(product_variations) > 0:
                    digikey_pn = product_variations[0].get('DigiKeyProductNumber', '')
                
                # Fetch pricing from Product Pricing API if we have a DigiKey P/N
                pricing_data = None
                if digikey_pn:
                    self.log(f"  💰 Fetching pricing for {digikey_pn}...")
                    pricing_data = self.dk.get_product_pricing(digikey_pn)
                    if pricing_data:
                        self.log(f"  ✅ Pricing data retrieved")
                    else:
                        self.log(f"  ⚠️  Could not fetch pricing, using search results")
                    time.sleep(delay_between_requests)  # Additional delay for pricing API call
                
                data = self._extract_product_data(product, pricing_data)
                
                # Populate cells
                self._populate_row(sheet, row_idx, column_map, data)
                
                # Define common styles - Arial, size 10, black text
                arial_font = Font(name='Arial', size=10, color='000000')
                center_alignment = Alignment(horizontal='center', vertical='center')
                currency_format = '"$"#,##0.00'
                
                # Color green for success and apply font/alignment to all cells
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.fill = green_fill
                    # Preserve hyperlink font styling for dist_pn column
                    if 'dist_pn' in column_map and col == column_map['dist_pn'] and data['dist_pn']:
                        cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')
                    else:
                        cell.font = arial_font
                    cell.alignment = center_alignment
                
                # Apply currency formatting to price columns
                if 'price' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price']).number_format = currency_format
                if 'price_total' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price_total']).number_format = currency_format
                
                self.log(f"  ✅ SUCCESS")
                self.log(f"     DigiKey P/N: {data['dist_pn'] or '(not available)'}")
                self.log(f"     Manufacturer: {data['manufacturer'] or '(not available)'}")
                self.log(f"     Stock: {data['available']}")
                self.log(f"     Price: ${data['price']}")
                self.log(f"     Temp: {data['temperature'] or '(not available)'}")
                
                stats['success'] += 1
                
            else:
                # Define common styles - Arial, size 10, black text
                arial_font = Font(name='Arial', size=10, color='000000')
                center_alignment = Alignment(horizontal='center', vertical='center')
                currency_format = '"$"#,##0.00'
                
                # Not found - color red and apply font/alignment to all cells
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.fill = red_fill
                    cell.font = arial_font
                    cell.alignment = center_alignment
                
                # Apply currency formatting to price columns even for not found
                if 'price' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price']).number_format = currency_format
                if 'price_total' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price_total']).number_format = currency_format
                
                if 'distributor' in column_map:
                    sheet.cell(row=row_idx, column=column_map['distributor']).value = "NOT FOUND"
                
                self.log(f"  ❌ NOT FOUND")
                stats['failed'] += 1
                stats['failed_parts'].append(mfr_pn)
            
            # Delay between requests
            time.sleep(delay_between_requests)
            
            # Save progress every 10 parts
            if stats['processed'] % 10 == 0:
                self.log(f"\n💾 Saving progress... ({stats['success']}/{stats['processed']} successful)")
                try:
                    temp_file = output_file or f"{os.path.splitext(bom_file_path)[0]}_temp.xlsx"
                    workbook.save(temp_file)
                except Exception as e:
                    self.log(f"⚠️  Could not save progress: {e}")
        
        # Save final result
        if output_file is None:
            output_file = f"{os.path.splitext(bom_file_path)[0]}_populated.xlsx"
        
        self.log("\n" + "="*70)
        self.log("💾 Saving final BOM...")
        try:
            workbook.save(output_file)
            self.log(f"✅ Saved to: {output_file}")
        except Exception as e:
            self.log(f"❌ Error saving: {e}")
            return
        
        # Print summary
        self._print_summary(stats, output_file)
    
    def _find_columns(self, sheet):
        """Find header row and map columns"""
        header_row = None
        column_map = {}
        
        # Search first 10 rows
        for row_idx in range(1, 11):
            for col_idx, cell in enumerate(sheet[row_idx], start=1):
                cell_value = str(cell.value).strip() if cell.value else ""
                
                if "Mfr" in cell_value and "P/N" in cell_value:
                    header_row = row_idx
                    break
            if header_row:
                break
        
        if not header_row:
            return None, {}
        
        # Map columns
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            cell_value = str(cell.value).strip() if cell.value else ""
            
            if "Mfr" in cell_value and "P/N" in cell_value:
                column_map['mfr_pn'] = col_idx
            elif cell_value == "Mfr." or cell_value == "Mfr" or cell_value == "Manufacturer":
                column_map['manufacturer'] = col_idx
            elif "Temperature" in cell_value:
                column_map['temperature'] = col_idx
            elif "Description" in cell_value:
                column_map['description'] = col_idx
            elif "Distribut" in cell_value:
                column_map['distributor'] = col_idx
            elif "Dist" in cell_value and "P/N" in cell_value:
                column_map['dist_pn'] = col_idx
            elif cell_value == "Price Ea" or cell_value == "Price Each" or cell_value == "Unit Price":
                column_map['price'] = col_idx
            elif "Price Total Per Board" in cell_value:
                column_map['price_total'] = col_idx
            elif "Date Updated" in cell_value:
                column_map['date_updated'] = col_idx
            elif "Available" in cell_value:
                column_map['available'] = col_idx
        
        return header_row, column_map
    
    def _extract_product_data(self, product: Dict, pricing_data: Optional[Dict] = None) -> Dict:
        """Extract relevant data from DigiKey product"""
        # Get UnitPrice directly from product data
        # First try pricing_data if available, then fall back to product data
        unit_price = 0
        if pricing_data:
            # Try to get UnitPrice directly from pricing API response
            unit_price = pricing_data.get('UnitPrice', 0)
        
        # Fall back to product data if pricing API didn't return UnitPrice
        if unit_price == 0:
            unit_price = product.get('UnitPrice', 0)
        
        # Extract temperature - look for "Operating Temperature" in ParameterText
        # First try pricing_data (ProductDetails) as it has more complete parameters
        temperature = ""
        parameters = []
        if pricing_data:
            parameters = pricing_data.get('Parameters', [])
        if not parameters:
            parameters = product.get('Parameters', [])
        
        for param in parameters:
            param_text = param.get('ParameterText', '').lower()
            # Check for various temperature-related parameter names
            if 'operating temperature' in param_text or 'temperature - operating' in param_text:
                temp_value = param.get('ValueText', '')
                if temp_value and temp_value != 'N/A' and temp_value != '-':
                    temperature = temp_value
                    break
        
        # Handle description (can be string or dict)
        description_raw = product.get('Description', 'N/A')
        if isinstance(description_raw, dict):
            description = description_raw.get('ProductDescription') or \
                         description_raw.get('DetailedDescription') or \
                         'N/A'
        else:
            description = description_raw
        
        # Get DigiKey Product Number - it's inside ProductVariations array
        digikey_pn = ''
        # First try pricing_data (ProductDetails response) which may have it at the top level
        if pricing_data:
            digikey_pn = pricing_data.get('DigiKeyProductNumber', '')
        
        # If not found, try ProductVariations array from keyword search
        if not digikey_pn:
            product_variations = product.get('ProductVariations', [])
            if product_variations and len(product_variations) > 0:
                digikey_pn = product_variations[0].get('DigiKeyProductNumber', '')
        
        # Also check top level just in case (for ProductDetails endpoint)
        if not digikey_pn:
            digikey_pn = product.get('DigiKeyProductNumber', '')
        
        if digikey_pn == 'N/A':
            digikey_pn = ''  # Leave blank instead of N/A
        
        # Get Manufacturer Name from Manufacturer dictionary
        manufacturer = ''
        manufacturer_data = product.get('Manufacturer', {})
        if isinstance(manufacturer_data, dict):
            manufacturer = manufacturer_data.get('Name', '')
        
        return {
            'description': description,
            'dist_pn': digikey_pn,
            'product_url': product.get('ProductUrl', ''),
            'available': product.get('QuantityAvailable', 0),
            'price': unit_price,
            'temperature': temperature,
            'distributor': 'DigiKey',
            'manufacturer': manufacturer
        }
    
    def _populate_row(self, sheet, row_idx: int, column_map: Dict, data: Dict):
        """Populate a single row with data"""
        # Define common styles - Arial, size 10, black text
        arial_font = Font(name='Arial', size=10, color='000000')
        center_alignment = Alignment(horizontal='center', vertical='center')
        currency_format = '"$"#,##0.00'
        
        if 'description' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['description'])
            cell.value = data['description']
            cell.font = arial_font
            cell.alignment = center_alignment
        
        if 'temperature' in column_map and data['temperature']:
            cell = sheet.cell(row=row_idx, column=column_map['temperature'])
            cell.value = data['temperature']
            cell.font = arial_font
            cell.alignment = center_alignment
        
        if 'distributor' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['distributor'])
            cell.value = data['distributor']
            cell.font = arial_font
            cell.alignment = center_alignment
        
        if 'manufacturer' in column_map and data.get('manufacturer'):
            cell = sheet.cell(row=row_idx, column=column_map['manufacturer'])
            cell.value = data['manufacturer']
            cell.font = arial_font
            cell.alignment = center_alignment
        
        if 'dist_pn' in column_map and data['dist_pn']:
            cell = sheet.cell(row=row_idx, column=column_map['dist_pn'])
            cell.value = data['dist_pn']
            cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')  # Hyperlink style
            cell.alignment = center_alignment
            # Add hyperlink if product URL exists
            if data.get('product_url'):
                cell.hyperlink = data['product_url']
        
        if 'price' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['price'])
            cell.value = data['price']
            cell.font = arial_font
            cell.alignment = center_alignment
            cell.number_format = currency_format
        
        if 'price_total' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['price_total'])
            # Keep existing value if any, just apply formatting
            cell.font = arial_font
            cell.alignment = center_alignment
            cell.number_format = currency_format
        
        if 'date_updated' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['date_updated'])
            cell.value = datetime.now().strftime('%m/%d/%Y')
            cell.font = arial_font
            cell.alignment = center_alignment
        
        if 'available' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['available'])
            cell.value = data['available']
            cell.font = arial_font
            cell.alignment = center_alignment
    
    def _print_summary(self, stats: Dict, output_file: str):
        """Print completion summary"""
        self.log("\n" + "="*70)
        self.log("📊 SUMMARY")
        self.log("="*70)
        self.log(f"Total parts processed: {stats['processed']}")
        self.log(f"Successfully found: {stats['success']}")
        self.log(f"Not found: {stats['failed']}")
        self.log(f"Skipped (empty): {stats['skipped']}")
        self.log(f"API requests made: {self.dk.request_count}")
        
        if stats['failed_parts']:
            self.log(f"\n⚠️  Parts not found ({len(stats['failed_parts'])}):")
            for part in stats['failed_parts']:
                self.log(f"   - {part}")
        
        success_rate = (stats['success'] / stats['processed'] * 100) if stats['processed'] > 0 else 0
        self.log(f"\n📈 Success rate: {success_rate:.1f}%")
        self.log(f"\n✅ BOM population complete!")
        self.log(f"📁 Output file: {output_file}")
        self.log(f"📋 Log file: {self.log_file}")
        self.log("="*70)


# ============================================================
# MAIN PROGRAM
# ============================================================

if __name__ == "__main__":
    
    # Configuration
    DIGIKEY_CLIENT_ID = "ENciL4CxjO0MAfsCMveTtNb2Ke96Xi5vSt5vanaouZIQbTq5"
    DIGIKEY_CLIENT_SECRET = "gGRJn44v69tnXdYUUEM27ELExG64yG4LoyQdVKegJlpDHTvDbXU87cevg6HvHIoo"
    
    BOM_INPUT_FILE = "Motherboard Panel Connectors.xlsx"
    BOM_OUTPUT_FILE = "Motherboard Panel Connectors_Populated.xlsx"
    
    # Delay between API requests (seconds)
    # Increase if you hit rate limits
    REQUEST_DELAY = 0.5
    
    # Run the populator
    populator = BOMPopulator(DIGIKEY_CLIENT_ID, DIGIKEY_CLIENT_SECRET)
    
    populator.populate_bom(
        bom_file_path=BOM_INPUT_FILE,
        output_file=BOM_OUTPUT_FILE,
        delay_between_requests=REQUEST_DELAY
    )