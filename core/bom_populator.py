import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from typing import Dict, Optional, Callable
import threading
import time
import os

from core.digikey_api import DigiKeyAPI
from core.mouser_api import MouserAPI
from core.newark_api import NewarkAPI

# Priority order for each distributor mode
DISTRIBUTOR_PRIORITY = {
    'DigiKey': ['DigiKey'],
    'Mouser': ['Mouser'],
    'Newark': ['Newark'],
    'DigiKey 1st': ['DigiKey', 'Mouser', 'Newark'],
    'Mouser 1st': ['Mouser', 'DigiKey', 'Newark'],
    'Newark 1st': ['Newark', 'Mouser', 'DigiKey'],
}


class BOMPopulator:
    """Main BOM population class with callback support for GUI."""

    def __init__(self, client_id: str = '', client_secret: str = '',
                 mouser_api_key: str = '', newark_api_key: str = '',
                 distributor: str = 'DigiKey',
                 log_callback: Callable[[str], None] = None,
                 progress_callback: Callable[[int, int, str, str], None] = None):
        self._log = log_callback or print
        self._progress = progress_callback
        self.distributor = distributor
        self.dk = DigiKeyAPI(client_id, client_secret, log_callback=self._log) if client_id else None
        self.mouser = MouserAPI(mouser_api_key, log_callback=self._log) if mouser_api_key else None
        self.newark = NewarkAPI(newark_api_key, log_callback=self._log) if newark_api_key else None
        self.log_file = None
        self._cancel_event = threading.Event()

    def cancel(self):
        self._cancel_event.set()

    def log(self, message: str):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._log(message)
        if self.log_file:
            try:
                with open(self.log_file, 'a', encoding='utf-8') as f:
                    f.write(f"[{timestamp}] {message}\n")
            except Exception:
                pass

    def populate_bom(self, bom_file_path: str = None, workbook: openpyxl.Workbook = None,
                     output_file: str = None, bom_name: str = None,
                     num_boards: int = 1,
                     delay_between_requests: float = 0.5) -> dict:
        self._cancel_event.clear()

        # Setup logging
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = f"bom_population_{timestamp}.log"

        self.log("=" * 70)
        self.log("BOM AUTO-POPULATION - DIGIKEY")
        self.log("=" * 70)

        # Load workbook
        if workbook:
            sheet = workbook.active
            self.log(f"Using pre-built workbook: {sheet.title}")
        elif bom_file_path:
            self.log(f"Input file: {bom_file_path}")
            try:
                workbook = openpyxl.load_workbook(bom_file_path)
                sheet = workbook.active
                self.log(f"Loaded sheet: {sheet.title}")
            except Exception as e:
                self.log(f"Error loading file: {e}")
                return {'processed': 0, 'success': 0, 'failed': 0, 'skipped': 0, 'failed_parts': []}
        else:
            self.log("Error: No input file or workbook provided")
            return {'processed': 0, 'success': 0, 'failed': 0, 'skipped': 0, 'failed_parts': []}

        # Update BOM name in rows 1-2 if provided, with date
        if bom_name:
            date_str = datetime.now().strftime('%B %d, %Y')
            sheet.cell(row=1, column=2).value = f"{bom_name}, {date_str}"

        # Find columns
        self.log("\nLocating columns...")
        header_row, column_map = self._find_columns(sheet)

        if not header_row or 'mfr_pn' not in column_map:
            self.log("Could not find required columns (need at least 'Mfr. P/N')")
            return {'processed': 0, 'success': 0, 'failed': 0, 'skipped': 0, 'failed_parts': []}

        self.log(f"Header at row {header_row}")
        self.log(f"Columns: {list(column_map.keys())}")

        # Apply borders to header row
        thin_border_h = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row, column=col)
            if cell.value:
                cell.border = thin_border_h

        # Determine which distributors to use based on priority
        priority = DISTRIBUTOR_PRIORITY.get(self.distributor, ['DigiKey'])
        active_distributors = []

        for dist in priority:
            if dist == 'DigiKey' and self.dk:
                self.log(f"\nAuthenticating with DigiKey...")
                if self.dk.authenticate():
                    self.log("DigiKey authentication successful")
                    active_distributors.append('DigiKey')
                else:
                    self.log("DigiKey authentication failed")
            elif dist == 'Mouser' and self.mouser:
                self.log("Mouser API ready")
                active_distributors.append('Mouser')
            elif dist == 'Newark' and self.newark:
                self.log("Newark API ready")
                active_distributors.append('Newark')

        if not active_distributors:
            self.log("No distributor API available")
            return {'processed': 0, 'success': 0, 'failed': 0, 'skipped': 0, 'failed_parts': []}
        self.log("Authentication successful")

        # Count total parts to process
        data_start_row = header_row + 1
        total_parts = 0
        for row_idx in range(data_start_row, sheet.max_row + 1):
            mfr_pn_cell = sheet.cell(row=row_idx, column=column_map['mfr_pn'])
            mfr_pn = str(mfr_pn_cell.value).strip() if mfr_pn_cell.value else ""
            if mfr_pn and mfr_pn.lower() not in ['none', 'n/a', '', 'null']:
                total_parts += 1

        self.log(f"\nProcessing {total_parts} parts...")
        self.log("-" * 70)

        stats = {
            'processed': 0, 'success': 0, 'failed': 0, 'skipped': 0,
            'failed_parts': []
        }

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        purple_fill = PatternFill(start_color="D9C4EC", end_color="D9C4EC", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        arial_font = Font(name='Arial', size=10, color='000000')
        center = Alignment(horizontal='center', vertical='center')
        currency_format = '"$"#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for row_idx in range(data_start_row, sheet.max_row + 1):
            if self._cancel_event.is_set():
                self.log("\nCancelled by user.")
                break

            mfr_pn_cell = sheet.cell(row=row_idx, column=column_map['mfr_pn'])
            mfr_pn = str(mfr_pn_cell.value).strip() if mfr_pn_cell.value else ""

            if not mfr_pn or mfr_pn.lower() in ['none', 'n/a', '', 'null']:
                stats['skipped'] += 1
                continue

            # Check for DNI in notes column - highlight red and skip lookup
            if 'notes' in column_map:
                notes_val = str(sheet.cell(row=row_idx, column=column_map['notes']).value or '').strip()
                if notes_val.upper().startswith('DNI'):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        cell.fill = red_fill
                        cell.font = arial_font
                        cell.alignment = center
                        cell.border = thin_border
                    stats['skipped'] += 1
                    self.log(f"\n  Skipping DNI part: {mfr_pn}")
                    continue

            stats['processed'] += 1
            self.log(f"\n[{stats['processed']}/{total_parts}] Part: {mfr_pn}")

            if self._progress:
                self._progress(stats['processed'], total_parts, mfr_pn, 'searching')

            # Search distributor(s) in priority order
            data = None
            for dist in active_distributors:
                if dist == 'DigiKey':
                    result = self._search_digikey(mfr_pn, delay_between_requests)
                elif dist == 'Mouser':
                    result = self._search_mouser(mfr_pn)
                    time.sleep(delay_between_requests)
                elif dist == 'Newark':
                    result = self._search_newark(mfr_pn)
                    time.sleep(delay_between_requests)
                else:
                    continue

                if result:
                    if not data:
                        data = result
                    # If current best has 0 available and this one has stock, switch
                    if data.get('available', 0) == 0 and result.get('available', 0) > 0:
                        self.log(f"  {data['distributor']} shows 0 available, using {result['distributor']}")
                        data = result
                    # If we found something with stock, stop searching
                    if data.get('available', 0) > 0:
                        break

            # Supplement missing detail fields (value, temperature, footprint)
            # from other distributors if the primary didn't provide them
            if data and len(active_distributors) > 1:
                missing_fields = []
                if not data.get('component_value'):
                    missing_fields.append('component_value')
                if not data.get('temperature'):
                    missing_fields.append('temperature')
                if not data.get('footprint'):
                    missing_fields.append('footprint')

                if missing_fields:
                    # Search other distributors for detail data
                    detail_sources = [d for d in active_distributors if d != data.get('distributor')]
                    for src in detail_sources:
                        if not missing_fields:
                            break
                        supplement = None
                        try:
                            if src == 'DigiKey' and self.dk:
                                self.log(f"  Fetching details from DigiKey...")
                                supplement = self._search_digikey(mfr_pn, delay_between_requests, details_only=True)
                            elif src == 'Newark' and self.newark:
                                self.log(f"  Fetching details from Newark...")
                                supplement = self._search_newark(mfr_pn)
                                time.sleep(delay_between_requests)
                        except Exception:
                            self.log(f"  {src} detail fetch failed, trying next...")
                            continue

                        if supplement:
                            for field in list(missing_fields):
                                if supplement.get(field):
                                    data[field] = supplement[field]
                                    missing_fields.remove(field)
                            if not missing_fields:
                                break
                        else:
                            self.log(f"  {src} returned no detail data, trying next...")

            if data:
                self._populate_row(sheet, row_idx, column_map, data,
                                   item_number=stats['processed'], num_boards=num_boards)

                # Determine row fill color based on which distributor found it
                zero_available = data['available'] == 0 or data['available'] is None
                used_dist = data.get('distributor', '')

                # Find position of the used distributor in priority list
                if used_dist in active_distributors:
                    dist_rank = active_distributors.index(used_dist)
                else:
                    dist_rank = 0

                if zero_available:
                    row_fill = red_fill
                elif dist_rank == 0:
                    row_fill = green_fill   # Primary distributor
                elif dist_rank == 1:
                    row_fill = blue_fill    # Secondary distributor
                else:
                    row_fill = purple_fill  # Tertiary distributor

                # Apply fill, font, alignment, borders to all cells
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.fill = row_fill
                    if 'dist_pn' in column_map and col == column_map['dist_pn'] and data['dist_pn']:
                        cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')
                    else:
                        cell.font = arial_font
                    cell.alignment = center
                    cell.border = thin_border

                if 'price' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price']).number_format = currency_format
                if 'price_total' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price_total']).number_format = currency_format

                # Handle notes column
                if 'notes' in column_map:
                    notes_cell = sheet.cell(row=row_idx, column=column_map['notes'])
                    existing = str(notes_cell.value).strip() if notes_cell.value else ""
                    # Preserve DNI, clear everything else
                    has_dni = existing.upper().startswith("DNI")
                    notes_parts = ["DNI"] if has_dni else []
                    if zero_available:
                        notes_parts.append("None available")
                    notes_cell.value = ", ".join(notes_parts) if notes_parts else ""

                if zero_available:
                    self.log(f"  SUCCESS (0 AVAILABLE) - {data['dist_pn'] or 'N/A'} | ${data['price']} | Stock: 0")
                else:
                    self.log(f"  SUCCESS - {data['dist_pn'] or 'N/A'} | ${data['price']} | Stock: {data['available']}")
                stats['success'] += 1

                if self._progress:
                    self._progress(stats['processed'], total_parts, mfr_pn, 'found')
            else:
                # Not found by any distributor - red fill with borders
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.fill = red_fill
                    cell.font = arial_font
                    cell.alignment = center
                    cell.border = thin_border

                if 'price' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price']).number_format = currency_format
                if 'price_total' in column_map:
                    sheet.cell(row=row_idx, column=column_map['price_total']).number_format = currency_format

                if 'distributor' in column_map:
                    sheet.cell(row=row_idx, column=column_map['distributor']).value = "NOT FOUND"

                # Item number for not-found rows
                if 'item' in column_map:
                    sheet.cell(row=row_idx, column=column_map['item']).value = stats['processed']

                # Number of Boards
                if 'num_boards' in column_map:
                    sheet.cell(row=row_idx, column=column_map['num_boards']).value = num_boards

                # Formulas for not-found rows too
                if 'qty_total' in column_map and 'qty_per_board' in column_map and 'num_boards' in column_map:
                    from openpyxl.utils import get_column_letter
                    qpb_col = get_column_letter(column_map['qty_per_board'])
                    nb_col = get_column_letter(column_map['num_boards'])
                    sheet.cell(row=row_idx, column=column_map['qty_total']).value = f"={qpb_col}{row_idx}*{nb_col}{row_idx}"

                if 'price_total' in column_map and 'qty_total' in column_map and 'price' in column_map:
                    from openpyxl.utils import get_column_letter
                    qt_col = get_column_letter(column_map['qty_total'])
                    pr_col = get_column_letter(column_map['price'])
                    cell = sheet.cell(row=row_idx, column=column_map['price_total'])
                    cell.value = f"={qt_col}{row_idx}*{pr_col}{row_idx}"
                    cell.number_format = currency_format

                if 'total_buy' in column_map and 'qty_to_buy' in column_map and 'price' in column_map:
                    from openpyxl.utils import get_column_letter
                    qtb_col = get_column_letter(column_map['qty_to_buy'])
                    pr_col = get_column_letter(column_map['price'])
                    cell = sheet.cell(row=row_idx, column=column_map['total_buy'])
                    cell.value = f"={qtb_col}{row_idx}*{pr_col}{row_idx}"
                    cell.number_format = currency_format

                # Handle notes for not found
                if 'notes' in column_map:
                    notes_cell = sheet.cell(row=row_idx, column=column_map['notes'])
                    existing = str(notes_cell.value).strip() if notes_cell.value else ""
                    has_dni = existing.upper().startswith("DNI")
                    notes_parts = ["DNI"] if has_dni else []
                    notes_parts.append("Not found on DigiKey")
                    notes_cell.value = ", ".join(notes_parts)

                self.log(f"  NOT FOUND")
                stats['failed'] += 1
                stats['failed_parts'].append(mfr_pn)

                if self._progress:
                    self._progress(stats['processed'], total_parts, mfr_pn, 'not_found')

            time.sleep(delay_between_requests)

            # Save progress every 10 parts
            if stats['processed'] % 10 == 0 and output_file:
                try:
                    workbook.save(output_file)
                    self.log(f"\nProgress saved ({stats['success']}/{stats['processed']} found)")
                except Exception:
                    pass

        # Add totals row underneath the last line item
        if stats['processed'] > 0:
            from openpyxl.utils import get_column_letter
            # Find the last row that was processed
            last_data_row = data_start_row
            for r in range(data_start_row, sheet.max_row + 1):
                mfr_val = sheet.cell(row=r, column=column_map['mfr_pn']).value
                if mfr_val and str(mfr_val).strip().lower() not in ('none', 'n/a', '', 'null'):
                    last_data_row = r

            totals_row = last_data_row + 2
            bold_font = Font(name='Arial', size=10, bold=True, color='000000')

            if 'price_total' in column_map:
                pt_col = get_column_letter(column_map['price_total'])
                cell = sheet.cell(row=totals_row, column=column_map['price_total'])
                cell.value = f"=SUM({pt_col}{data_start_row}:{pt_col}{last_data_row})"
                cell.font = bold_font
                cell.alignment = center
                cell.number_format = currency_format
                cell.border = thin_border
                # Label
                label_col = column_map['price_total'] - 1
                label_cell = sheet.cell(row=totals_row, column=label_col)
                label_cell.value = "TOTAL"
                label_cell.font = bold_font
                label_cell.alignment = center

            if 'total_buy' in column_map:
                tb_col = get_column_letter(column_map['total_buy'])
                cell = sheet.cell(row=totals_row, column=column_map['total_buy'])
                cell.value = f"=SUM({tb_col}{data_start_row}:{tb_col}{last_data_row})"
                cell.font = bold_font
                cell.alignment = center
                cell.number_format = currency_format
                cell.border = thin_border

        # Determine output filename
        if not output_file:
            if bom_name:
                safe_name = "".join(c for c in bom_name if c not in '<>:"/\\|?*')
                output_file = f"{safe_name}_Populated.xlsx"
            elif bom_file_path:
                output_file = f"{os.path.splitext(bom_file_path)[0]}_Populated.xlsx"
            else:
                output_file = "BOM_Populated.xlsx"

        # Save
        self.log("\n" + "=" * 70)
        try:
            workbook.save(output_file)
            self.log(f"Saved to: {output_file}")
        except Exception as e:
            self.log(f"Error saving: {e}")

        # Summary
        self.log("\n" + "=" * 70)
        self.log("SUMMARY")
        self.log("=" * 70)
        self.log(f"Total parts processed: {stats['processed']}")
        self.log(f"Successfully found: {stats['success']}")
        self.log(f"Not found: {stats['failed']}")
        self.log(f"Skipped (empty): {stats['skipped']}")
        self.log(f"API requests made: {self.dk.request_count}")

        if stats['failed_parts']:
            self.log(f"\nParts not found ({len(stats['failed_parts'])}):")
            for part in stats['failed_parts']:
                self.log(f"   - {part}")

        success_rate = (stats['success'] / stats['processed'] * 100) if stats['processed'] > 0 else 0
        self.log(f"\nSuccess rate: {success_rate:.1f}%")
        self.log(f"Output file: {output_file}")
        self.log("=" * 70)

        return stats

    @staticmethod
    def _get_bulk_passive_step(description: str) -> int:
        """Return the step size for bulk buying, or 0 if not a bulk passive.
        Resistors: step 50, Capacitors/Ferrite beads: step 5.
        """
        desc = (description or '').lower()
        # Resistors -> step 50
        resistor_kw = ['resistor', 'res ', 'res.', 'ohm', 'thick film', 'thin film']
        if any(kw in desc for kw in resistor_kw):
            return 50
        # Capacitors and ferrite beads -> step 5
        cap_ferrite_kw = ['capacitor', 'cap ', 'cap.', 'farad', 'ceramic', 'mlcc',
                          'ferrite', 'bead']
        if any(kw in desc for kw in cap_ferrite_kw):
            return 5
        return 0

    @staticmethod
    def _calculate_qty_to_buy(qty_total: int, price_breaks: list,
                              is_bulk: bool, max_budget: float = 25.0,
                              max_qty: int = 1000, step: int = 50) -> tuple:
        """Calculate optimal QTY TO BUY and the unit price at that quantity.
        Returns (qty_to_buy, unit_price_at_qty).
        """
        import math
        # Minimum: qty_total + 10%, rounded up
        min_qty = math.ceil(qty_total * 1.1)
        if min_qty < 1:
            min_qty = 1

        if not price_breaks or not is_bulk:
            # Non R/C parts: just buy minimum + 10%
            # Find the best price break for min_qty
            best_price = price_breaks[0]['unit_price'] if price_breaks else 0
            for pb in sorted(price_breaks, key=lambda x: x['quantity']):
                if pb['quantity'] <= min_qty:
                    best_price = pb['unit_price']
            return min_qty, best_price

        # R/C parts: try to buy more if it's cheap
        # Sort price breaks by quantity ascending
        breaks = sorted(price_breaks, key=lambda x: x['quantity'])

        def get_unit_price(qty):
            """Get the unit price for a given quantity based on price breaks."""
            price = breaks[0]['unit_price'] if breaks else 0
            for pb in breaks:
                if pb['quantity'] <= qty:
                    price = pb['unit_price']
                else:
                    break
            return price

        # If minimum needed already costs > budget, just buy the minimum
        min_cost = min_qty * get_unit_price(min_qty)
        if min_cost > max_budget:
            return min_qty, get_unit_price(min_qty)

        # If minimum needed > max_qty, just buy minimum
        if min_qty > max_qty:
            return min_qty, get_unit_price(min_qty)

        # Try stepping up from min_qty in steps of 'step' up to max_qty
        # Find the highest quantity where total cost <= budget
        best_qty = min_qty
        best_price = get_unit_price(min_qty)

        # Round min_qty up to next step
        start = max(min_qty, step)
        if start % step != 0:
            start = ((start // step) + 1) * step

        for qty in range(start, max_qty + 1, step):
            up = get_unit_price(qty)
            total = qty * up
            if total <= max_budget:
                best_qty = qty
                best_price = up
            else:
                break

        # Also check max_qty exactly
        up = get_unit_price(max_qty)
        if max_qty * up <= max_budget:
            best_qty = max_qty
            best_price = up

        # Make sure we're at least buying minimum
        if best_qty < min_qty:
            best_qty = min_qty
            best_price = get_unit_price(min_qty)

        return best_qty, best_price

    @staticmethod
    def _normalize_pn(pn: str) -> str:
        """Normalize a part number for comparison by stripping dashes, spaces, and leading zeros."""
        return pn.upper().replace('-', '').replace(' ', '').replace('.', '').lstrip('0')

    def _search_digikey(self, mfr_pn: str, delay: float = 0.5, details_only: bool = False) -> Optional[Dict]:
        """Search DigiKey and return extracted product data, or None.
        If details_only=True, skip the pricing API call (faster, for supplementing data).
        """
        result = self.dk.search_part(mfr_pn)
        if not result or not result.get('Products'):
            return None

        products = result['Products']
        product = None
        norm_search = self._normalize_pn(mfr_pn)

        # Exact match first (literal)
        for p in products:
            mpn = p.get('ManufacturerProductNumber', '')
            if mpn.upper() == mfr_pn.upper():
                product = p
                self.log(f"  [DigiKey] Exact match: {mpn}")
                break

        # Normalized match (strips dashes, leading zeros, etc.)
        if not product:
            for p in products:
                mpn = p.get('ManufacturerProductNumber', '')
                if self._normalize_pn(mpn) == norm_search:
                    product = p
                    self.log(f"  [DigiKey] Normalized match: {mpn}")
                    break

        # ExactMatches array
        if not product and result.get('ExactMatches'):
            for p in result['ExactMatches']:
                mpn = p.get('ManufacturerProductNumber', '')
                if mpn.upper() == mfr_pn.upper() or self._normalize_pn(mpn) == norm_search:
                    product = p
                    self.log(f"  [DigiKey] Exact match: {mpn}")
                    break

        # Partial match
        if not product:
            for p in products:
                mpn = p.get('ManufacturerProductNumber', '')
                if mfr_pn.upper() in mpn.upper() or mpn.upper() in mfr_pn.upper():
                    product = p
                    self.log(f"  [DigiKey] Partial match: {mpn}")
                    break

        # Fallback
        if not product:
            product = products[0]
            self.log(f"  [DigiKey] Using first result: {product.get('ManufacturerProductNumber', 'N/A')}")

        # Get pricing - Cut Tape > Tape & Reel > Digi-Reel hierarchy
        digikey_pn = ''
        product_variations = product.get('ProductVariations', [])
        if product_variations:
            for preferred in ['cut tape', 'tape & reel', 'digi-reel']:
                for var in product_variations:
                    pkg_type = var.get('PackageType', {})
                    pkg_name = ''
                    if isinstance(pkg_type, dict):
                        pkg_name = pkg_type.get('Name', '').lower()
                    elif isinstance(pkg_type, str):
                        pkg_name = pkg_type.lower()
                    if preferred in pkg_name:
                        digikey_pn = var.get('DigiKeyProductNumber', '')
                        break
                if digikey_pn:
                    break
            if not digikey_pn:
                digikey_pn = product_variations[0].get('DigiKeyProductNumber', '')

        pricing_data = None
        if digikey_pn and not details_only:
            self.log(f"  Fetching pricing for {digikey_pn}...")
            pricing_data = self.dk.get_product_pricing(digikey_pn)
            time.sleep(delay)

        return self._extract_product_data(product, pricing_data)

    def _search_mouser(self, mfr_pn: str) -> Optional[Dict]:
        """Search Mouser and return extracted product data, or None."""
        return self.mouser.find_best_match(mfr_pn)

    def _search_newark(self, mfr_pn: str) -> Optional[Dict]:
        """Search Newark and return extracted product data, or None."""
        return self.newark.find_best_match(mfr_pn)

    def _find_columns(self, sheet):
        header_row = None
        column_map = {}

        for row_idx in range(1, 11):
            for col_idx, cell in enumerate(sheet[row_idx], start=1):
                cell_value = str(cell.value).strip() if cell.value else ""
                if "Mfr" in cell_value and "P/N" in cell_value:
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            return None, {}

        for col_idx, cell in enumerate(sheet[header_row], start=1):
            cell_value = str(cell.value).strip() if cell.value else ""

            if "Mfr" in cell_value and "P/N" in cell_value:
                column_map['mfr_pn'] = col_idx
            elif cell_value in ("Mfr.", "Mfr", "Manufacturer"):
                column_map['manufacturer'] = col_idx
            elif "Temperature" in cell_value:
                column_map['temperature'] = col_idx
            elif cell_value == "Value":
                column_map['value'] = col_idx
            elif "Description" in cell_value:
                column_map['description'] = col_idx
            elif "Distribut" in cell_value:
                column_map['distributor'] = col_idx
            elif "Dist" in cell_value and "P/N" in cell_value:
                column_map['dist_pn'] = col_idx
            elif cell_value in ("Price Ea", "Price Each", "Unit Price"):
                column_map['price'] = col_idx
            elif "Price Total Per Board" in cell_value:
                column_map['price_total'] = col_idx
            elif "Date Updated" in cell_value:
                column_map['date_updated'] = col_idx
            elif "Available" in cell_value:
                column_map['available'] = col_idx
            elif "Foot" in cell_value and "Print" in cell_value:
                column_map['footprint'] = col_idx
            elif cell_value == "Notes":
                column_map['notes'] = col_idx
            elif cell_value == "Item":
                column_map['item'] = col_idx
            elif "Quantity Per Board" in cell_value or "Number Per Board" in cell_value:
                column_map['qty_per_board'] = col_idx
            elif "Number of Boards" in cell_value:
                column_map['num_boards'] = col_idx
            elif "Quantity Total" in cell_value:
                column_map['qty_total'] = col_idx
            elif "Total Buy" in cell_value and "Price" not in cell_value:
                column_map['total_buy'] = col_idx
            elif "QTY TO BUY" in cell_value.upper():
                column_map['qty_to_buy'] = col_idx

        return header_row, column_map

    def _extract_product_data(self, product: Dict, pricing_data: Optional[Dict] = None) -> Dict:
        # Price
        unit_price = 0
        if pricing_data:
            unit_price = pricing_data.get('UnitPrice', 0)
        if unit_price == 0:
            unit_price = product.get('UnitPrice', 0)

        # Temperature
        temperature = ""
        parameters = []
        if pricing_data:
            parameters = pricing_data.get('Parameters', [])
        if not parameters:
            parameters = product.get('Parameters', [])

        for param in parameters:
            param_text = param.get('ParameterText', '').lower()
            if 'operating temperature' in param_text or 'temperature - operating' in param_text:
                temp_value = param.get('ValueText', '')
                if temp_value and temp_value not in ('N/A', '-'):
                    temperature = temp_value
                    break

        # Package / Footprint
        footprint = ""
        for param in parameters:
            param_text = param.get('ParameterText', '').lower()
            if 'package' in param_text or 'case' in param_text or 'footprint' in param_text:
                fp_value = param.get('ValueText', '')
                if fp_value and fp_value not in ('N/A', '-'):
                    footprint = fp_value
                    break

        # Component value (resistance, inductance, capacitance, etc.)
        component_value = ""
        value_params = [
            'resistance', 'inductance', 'capacitance', 'voltage - rated',
            'current rating', 'frequency',
        ]
        for param in parameters:
            param_text = param.get('ParameterText', '').lower()
            # Match primary value params but skip compound ones like "DC Resistance (DCR)"
            for vp in value_params:
                if param_text == vp or param_text.startswith(vp):
                    val = param.get('ValueText', '')
                    if val and val not in ('N/A', '-'):
                        component_value = val
                        break
            if component_value:
                break

        # Description
        description_raw = product.get('Description', 'N/A')
        if isinstance(description_raw, dict):
            description = description_raw.get('ProductDescription') or \
                         description_raw.get('DetailedDescription') or 'N/A'
        else:
            description = description_raw

        # DigiKey P/N - prioritize packaging: Cut Tape > Tape & Reel > Digi-Reel > first available
        digikey_pn = ''
        product_variations = product.get('ProductVariations', [])
        if product_variations:
            # Build a map of packaging type -> DigiKey P/N
            packaging_map = {}
            for var in product_variations:
                dk_pn = var.get('DigiKeyProductNumber', '')
                pkg_type = var.get('PackageType', {})
                pkg_name = ''
                if isinstance(pkg_type, dict):
                    pkg_name = pkg_type.get('Name', '').lower()
                elif isinstance(pkg_type, str):
                    pkg_name = pkg_type.lower()
                if dk_pn:
                    packaging_map[pkg_name] = dk_pn

            # Select by hierarchy: Cut Tape > Tape & Reel > Digi-Reel > first
            for preferred in ['cut tape', 'tape & reel', 'digi-reel']:
                for pkg_name, dk_pn in packaging_map.items():
                    if preferred in pkg_name:
                        digikey_pn = dk_pn
                        break
                if digikey_pn:
                    break

            # Fallback to first variation if no preferred packaging found
            if not digikey_pn:
                digikey_pn = product_variations[0].get('DigiKeyProductNumber', '')

        if not digikey_pn:
            digikey_pn = product.get('DigiKeyProductNumber', '')
        if digikey_pn == 'N/A':
            digikey_pn = ''

        # Manufacturer
        manufacturer = ''
        manufacturer_data = product.get('Manufacturer', {})
        if isinstance(manufacturer_data, dict):
            manufacturer = manufacturer_data.get('Name', '')

        # Extract price breaks from pricing data (ProductVariations -> StandardPricing)
        price_breaks = []
        if pricing_data:
            pd_product = pricing_data.get('Product', {})
            pd_variations = pd_product.get('ProductVariations', []) if pd_product else []
            # Find the variation matching our selected digikey_pn
            for var in pd_variations:
                if var.get('DigiKeyProductNumber', '') == digikey_pn:
                    for pb in var.get('StandardPricing', []):
                        price_breaks.append({
                            'quantity': pb.get('BreakQuantity', 0),
                            'unit_price': pb.get('UnitPrice', 0),
                        })
                    break
            # Fallback: use first variation's pricing
            if not price_breaks and pd_variations:
                for pb in pd_variations[0].get('StandardPricing', []):
                    price_breaks.append({
                        'quantity': pb.get('BreakQuantity', 0),
                        'unit_price': pb.get('UnitPrice', 0),
                    })

        return {
            'description': description,
            'dist_pn': digikey_pn,
            'product_url': product.get('ProductUrl', ''),
            'available': product.get('QuantityAvailable', 0),
            'price': unit_price,
            'price_breaks': price_breaks,
            'temperature': temperature,
            'footprint': footprint,
            'component_value': component_value,
            'distributor': 'DigiKey',
            'manufacturer': manufacturer
        }

    def _populate_row(self, sheet, row_idx: int, column_map: Dict, data: Dict,
                      item_number: int = None, num_boards: int = 1):
        arial_font = Font(name='Arial', size=10, color='000000')
        center = Alignment(horizontal='center', vertical='center')
        currency_format = '"$"#,##0.00'

        # Item number
        if 'item' in column_map and item_number is not None:
            cell = sheet.cell(row=row_idx, column=column_map['item'])
            cell.value = item_number
            cell.font = arial_font
            cell.alignment = center

        if 'description' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['description'])
            cell.value = data['description']
            cell.font = arial_font
            cell.alignment = center

        if 'temperature' in column_map and data['temperature']:
            cell = sheet.cell(row=row_idx, column=column_map['temperature'])
            cell.value = data['temperature']
            cell.font = arial_font
            cell.alignment = center

        if 'distributor' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['distributor'])
            cell.value = data['distributor']
            cell.font = arial_font
            cell.alignment = center

        if 'manufacturer' in column_map and data.get('manufacturer'):
            cell = sheet.cell(row=row_idx, column=column_map['manufacturer'])
            cell.value = data['manufacturer']
            cell.font = arial_font
            cell.alignment = center

        if 'dist_pn' in column_map and data['dist_pn']:
            cell = sheet.cell(row=row_idx, column=column_map['dist_pn'])
            cell.value = data['dist_pn']
            cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')
            cell.alignment = center
            if data.get('product_url'):
                cell.hyperlink = data['product_url']

        if 'price' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['price'])
            cell.value = data['price']
            cell.font = arial_font
            cell.alignment = center
            cell.number_format = currency_format

        # Number of Boards
        if 'num_boards' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['num_boards'])
            cell.value = num_boards
            cell.font = arial_font
            cell.alignment = center

        # Quantity Total = Qty Per Board * Number of Boards (formula referencing cells)
        if 'qty_total' in column_map and 'qty_per_board' in column_map and 'num_boards' in column_map:
            from openpyxl.utils import get_column_letter
            qpb_col = get_column_letter(column_map['qty_per_board'])
            nb_col = get_column_letter(column_map['num_boards'])
            cell = sheet.cell(row=row_idx, column=column_map['qty_total'])
            cell.value = f"={qpb_col}{row_idx}*{nb_col}{row_idx}"
            cell.font = arial_font
            cell.alignment = center

        # Price Total Per Board = Quantity Total * Price Ea (formula)
        if 'price_total' in column_map and 'qty_total' in column_map and 'price' in column_map:
            from openpyxl.utils import get_column_letter
            qt_col = get_column_letter(column_map['qty_total'])
            pr_col = get_column_letter(column_map['price'])
            cell = sheet.cell(row=row_idx, column=column_map['price_total'])
            cell.value = f"={qt_col}{row_idx}*{pr_col}{row_idx}"
            cell.font = arial_font
            cell.alignment = center
            cell.number_format = currency_format

        # Total Buy = QTY TO BUY * Price Ea (formula)
        if 'total_buy' in column_map and 'qty_to_buy' in column_map and 'price' in column_map:
            from openpyxl.utils import get_column_letter
            qtb_col = get_column_letter(column_map['qty_to_buy'])
            pr_col = get_column_letter(column_map['price'])
            cell = sheet.cell(row=row_idx, column=column_map['total_buy'])
            cell.value = f"={qtb_col}{row_idx}*{pr_col}{row_idx}"
            cell.font = arial_font
            cell.alignment = center
            cell.number_format = currency_format

        if 'date_updated' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['date_updated'])
            cell.value = datetime.now().strftime('%m/%d/%Y')
            cell.font = arial_font
            cell.alignment = center

        if 'available' in column_map:
            cell = sheet.cell(row=row_idx, column=column_map['available'])
            cell.value = data['available']
            cell.font = arial_font
            cell.alignment = center

        if 'footprint' in column_map and data.get('footprint'):
            cell = sheet.cell(row=row_idx, column=column_map['footprint'])
            cell.value = data['footprint']
            cell.font = arial_font
            cell.alignment = center

        # Component value (resistance, inductance, capacitance) - only if cell is empty
        if 'value' in column_map and data.get('component_value'):
            cell = sheet.cell(row=row_idx, column=column_map['value'])
            if not cell.value:
                cell.value = data['component_value']
                cell.font = arial_font
                cell.alignment = center

        # QTY TO BUY calculation
        if 'qty_to_buy' in column_map:
            # Get quantity total (try reading qty_per_board * num_boards directly)
            qty_total = 0
            if 'qty_per_board' in column_map:
                qpb_val = sheet.cell(row=row_idx, column=column_map['qty_per_board']).value
                try:
                    qty_total = int(qpb_val) * num_boards if qpb_val else 0
                except (ValueError, TypeError):
                    qty_total = 0

            if qty_total > 0:
                bulk_step = self._get_bulk_passive_step(data.get('description', ''))
                is_bulk = bulk_step > 0
                price_breaks = data.get('price_breaks', [])
                buy_qty, buy_price = self._calculate_qty_to_buy(
                    qty_total, price_breaks, is_bulk, step=bulk_step if is_bulk else 50)

                cell = sheet.cell(row=row_idx, column=column_map['qty_to_buy'])
                cell.value = buy_qty
                cell.font = arial_font
                cell.alignment = center

                # Update Price Ea if we got a better price at this quantity
                if buy_price > 0 and 'price' in column_map:
                    cell = sheet.cell(row=row_idx, column=column_map['price'])
                    cell.value = buy_price
                    cell.font = arial_font
                    cell.alignment = center
                    cell.number_format = currency_format
