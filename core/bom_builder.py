import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Standard 22-column BOM layout (columns B through V, A is spacer)
HEADERS = [
    'Item', 'Designator', 'Description', 'Temperature', 'Value',
    'Mfr.', 'Mfr. P/N', 'Quantity Per Board', 'Number of Boards',
    'Quantity Total', 'Distributor', 'Dist. P/N', 'Price Ea',
    'Price Total Per Board', 'Available', 'QTY TO BUY', 'Total Buy',
    'Date Updated', 'Foot Print', 'Notes'
]

# Maps part dict keys to header column indices (0-based offset from col B)
FIELD_TO_COL = {
    'designator': 1,    # Designator
    'description': 2,   # Description
    'value': 4,         # Value
    'mfr_pn': 6,        # Mfr. P/N
    'quantity': 7,       # Quantity Per Board
    'footprint': 18,     # Foot Print
    'comment': 19,       # Notes
}


def create_workbook(bom_name: str, parts: list, num_boards: int = 1) -> openpyxl.Workbook:
    """Create a 22-column BOM workbook from a list of part dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bom_name[:31] if len(bom_name) > 31 else bom_name

    header_font = Font(name='Arial', size=10, bold=True, color='000000')
    title_font = Font(name='Arial', size=14, bold=True, color='000000')
    data_font = Font(name='Arial', size=10, color='000000')
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Row 1-2: Merged title
    ws.merge_cells('B1:U2')
    title_cell = ws.cell(row=1, column=2)
    title_cell.value = bom_name
    title_cell.font = title_font
    title_cell.alignment = center

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Row 3: Headers (B3 through V3)
    for i, header in enumerate(HEADERS):
        cell = ws.cell(row=3, column=i + 2)
        cell.value = header
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    # Row 4: Empty (spacer)

    # Column letters for reference:
    # B=Item(2), C=Designator(3), D=Description(4), E=Type(5), F=Temperature(6),
    # G=Value(7), H=Mfr.(8), I=Mfr.P/N(9), J=QtyPerBoard(10), K=NumBoards(11),
    # L=QtyTotal(12), M=Distributor(13), N=Dist.P/N(14), O=PriceEa(15),
    # P=PriceTotalPerBoard(16), Q=Available(17), R=QtyToBuy(18), S=TotalBuy(19),
    # T=DateUpdated(20), U=FootPrint(21), V=Notes(22)

    currency_format = '"$"#,##0.00'

    # Row 5+: Data
    for idx, part in enumerate(parts):
        row = idx + 5
        # Item number (sequential)
        ws.cell(row=row, column=2).value = idx + 1

        # Apply borders, font, alignment to all cells in the row (B through U)
        for col in range(2, 22):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = data_font
            cell.alignment = center

        # Populate fields from part dict
        for field, col_offset in FIELD_TO_COL.items():
            value = part.get(field)
            if value is not None and str(value).strip():
                cell = ws.cell(row=row, column=col_offset + 2)
                if field == 'quantity':
                    try:
                        cell.value = int(value)
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = value

        # Number of Boards (col J = 10)
        ws.cell(row=row, column=10).value = num_boards

        # Quantity Total = Quantity Per Board * Number of Boards (col K = I * J)
        ws.cell(row=row, column=11).value = f"=I{row}*J{row}"

        # Price Total Per Board = Quantity Total * Price Ea (col O = K * N)
        price_cell = ws.cell(row=row, column=15)
        price_cell.value = f"=K{row}*N{row}"
        price_cell.number_format = currency_format

        # Price Ea formatting
        ws.cell(row=row, column=14).number_format = currency_format

        # Total Buy = QTY TO BUY * Price Ea (col R = Q * N)
        total_buy_cell = ws.cell(row=row, column=18)
        total_buy_cell.value = f"=Q{row}*N{row}"
        total_buy_cell.number_format = currency_format

    # Set reasonable column widths
    col_widths = {
        'A': 3, 'B': 6, 'C': 20, 'D': 35, 'E': 15, 'F': 10,
        'G': 15, 'H': 22, 'I': 14, 'J': 14, 'K': 14, 'L': 14,
        'M': 22, 'N': 10, 'O': 18, 'P': 12, 'Q': 12, 'R': 12,
        'S': 14, 'T': 14, 'U': 20
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    return wb
