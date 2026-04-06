"""
BOM Auto-Population Script for DigiKey
======================================

This script reads an Excel BOM file, searches DigiKey for each part,
and populates the fields automatically.

Requirements:
    pip install openpyxl requests --break-system-packages

Usage:
    python populate_bom.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import requests
import json
import time
from typing import Dict, Optional
import os


class DigiKeyAPI:
    """DigiKey API wrapper - Updated for API v4"""
    
    def __init__(self, client_id: str, client_secret: str):
        self.client_id = client_id
        self.client_secret = client_secret
        self.base_url = "https://api.digikey.com"
        self.access_token = None
        
    def authenticate(self):
        """Get OAuth2 access token"""
        auth_url = "https://api.digikey.com/v1/oauth2/token"
        
        data = {
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'grant_type': 'client_credentials'
        }
        
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        
        try:
            response = requests.post(auth_url, data=data, headers=headers, timeout=10)
            
            if response.status_code == 200:
                self.access_token = response.json()['access_token']
                print("✅ DigiKey authentication successful!")
                return True
            else:
                print(f"❌ Authentication failed: {response.status_code}")
                print(f"Response: {response.text}")
                return False
        except Exception as e:
            print(f"❌ Authentication error: {e}")
            return False
    
    def search_part(self, part_number: str) -> Optional[Dict]:
        """Search for a part by part number using API v4"""
        if not self.access_token:
            if not self.authenticate():
                return None
        
        url = f"{self.base_url}/products/v4/search/keyword"
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'X-DIGIKEY-Client-Id': self.client_id,
            'Content-Type': 'application/json'
        }
        
        payload = {
            'Keywords': part_number,
            'Limit': 10,
            'Offset': 0
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            
            if response.status_code == 200:
                return response.json()
            else:
                print(f"  ⚠️  Search failed for {part_number}: HTTP {response.status_code}")
                return None
        except Exception as e:
            print(f"  ⚠️  Search error for {part_number}: {e}")
            return None


def extract_temperature_from_parameters(parameters):
    """Extract operating temperature from product parameters"""
    if not parameters:
        return "N/A"
    
    # Look for temperature-related parameters
    temp_keywords = ['operating temperature', 'temperature range', 'temperature - operating']
    
    for param in parameters:
        param_name = param.get('Parameter', '').lower()
        
        for keyword in temp_keywords:
            if keyword in param_name:
                return param.get('Value', 'N/A')
    
    return "N/A"


def populate_bom_from_digikey(bom_file_path: str, client_id: str, client_secret: str, output_file: str = None):
    """
    Read BOM Excel file, search DigiKey, and populate fields
    
    Args:
        bom_file_path: Path to the input Excel BOM file
        client_id: DigiKey API client ID
        client_secret: DigiKey API client secret
        output_file: Path for output file (if None, creates one based on input filename)
    """
    
    print("="*70)
    print("BOM AUTO-POPULATION SCRIPT - DIGIKEY")
    print("="*70)
    
    # Load the Excel file
    print(f"\n📂 Loading BOM file: {bom_file_path}")
    try:
        workbook = openpyxl.load_workbook(bom_file_path)
        sheet = workbook.active
        print(f"✅ Loaded sheet: {sheet.title}")
    except Exception as e:
        print(f"❌ Error loading file: {e}")
        return
    
    # Find the header row and column indices
    print("\n🔍 Locating columns...")
    header_row = None
    column_map = {}
    
    # Search first 10 rows for headers
    for row_idx in range(1, 11):
        for col_idx, cell in enumerate(sheet[row_idx], start=1):
            cell_value = str(cell.value).strip() if cell.value else ""
            
            # Check if this looks like a header row
            if "Mfr. P/N" in cell_value or "Mfr P/N" in cell_value:
                header_row = row_idx
                break
        if header_row:
            break
    
    if not header_row:
        print("❌ Could not find 'Mfr. P/N' column. Please check your BOM format.")
        return
    
    print(f"✅ Found header row at row {header_row}")
    
    # Map column names to indices
    for col_idx, cell in enumerate(sheet[header_row], start=1):
        cell_value = str(cell.value).strip() if cell.value else ""
        
        if "Mfr. P/N" in cell_value or "Mfr P/N" in cell_value:
            column_map['mfr_pn'] = col_idx
        elif "Temperature" in cell_value:
            column_map['temperature'] = col_idx
        elif "Description" in cell_value:
            column_map['description'] = col_idx
        elif "Distributer" in cell_value or "Distributor" in cell_value:
            column_map['distributor'] = col_idx
        elif "Dist. P/N" in cell_value or "Dist P/N" in cell_value:
            column_map['dist_pn'] = col_idx
        elif "price of each" in cell_value.lower() or "unit price" in cell_value.lower():
            column_map['price'] = col_idx
        elif "Available" in cell_value:
            column_map['available'] = col_idx
    
    print(f"✅ Column mapping: {column_map}")
    
    if 'mfr_pn' not in column_map:
        print("❌ Could not find 'Mfr. P/N' column")
        return
    
    # Initialize DigiKey API
    print("\n🔐 Initializing DigiKey API...")
    dk = DigiKeyAPI(client_id, client_secret)
    
    # Process each row
    print("\n🔄 Processing BOM entries...")
    print("-"*70)
    
    data_start_row = header_row + 1
    processed_count = 0
    success_count = 0
    failed_parts = []
    
    # Color coding for status
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    for row_idx in range(data_start_row, sheet.max_row + 1):
        mfr_pn_cell = sheet.cell(row=row_idx, column=column_map['mfr_pn'])
        mfr_pn = str(mfr_pn_cell.value).strip() if mfr_pn_cell.value else ""
        
        # Skip empty rows
        if not mfr_pn or mfr_pn.lower() in ['none', 'n/a', '']:
            continue
        
        processed_count += 1
        print(f"\n[{processed_count}] Searching: {mfr_pn}")
        
        # Search DigiKey
        result = dk.search_part(mfr_pn)
        
        if result and result.get('Products'):
            product = result['Products'][0]  # Get first match
            
            # Extract data
            description = product.get('Description', 'N/A')
            digikey_pn = product.get('DigiKeyPartNumber', 'N/A')
            stock = product.get('QuantityAvailable', 0)
            
            # Get price (unit price for quantity 1)
            pricing = product.get('StandardPricing', [])
            unit_price = pricing[0].get('UnitPrice', 0) if pricing else 0
            
            # Extract temperature from parameters
            parameters = product.get('Parameters', [])
            temperature = extract_temperature_from_parameters(parameters)
            
            # Populate the BOM
            if 'description' in column_map:
                sheet.cell(row=row_idx, column=column_map['description']).value = description
            
            if 'temperature' in column_map:
                sheet.cell(row=row_idx, column=column_map['temperature']).value = temperature
            
            if 'distributor' in column_map:
                sheet.cell(row=row_idx, column=column_map['distributor']).value = "DigiKey"
            
            if 'dist_pn' in column_map:
                sheet.cell(row=row_idx, column=column_map['dist_pn']).value = digikey_pn
            
            if 'price' in column_map:
                sheet.cell(row=row_idx, column=column_map['price']).value = unit_price
            
            if 'available' in column_map:
                sheet.cell(row=row_idx, column=column_map['available']).value = stock
            
            # Color code the row (green = success)
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col).fill = green_fill
            
            print(f"  ✅ Found!")
            print(f"     Description: {description[:60]}...")
            print(f"     DigiKey P/N: {digikey_pn}")
            print(f"     Stock: {stock}")
            print(f"     Price: ${unit_price}")
            print(f"     Temperature: {temperature}")
            
            success_count += 1
            
        else:
            # Part not found - mark in red
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col).fill = red_fill
            
            if 'distributor' in column_map:
                sheet.cell(row=row_idx, column=column_map['distributor']).value = "NOT FOUND"
            
            print(f"  ❌ Part not found in DigiKey")
            failed_parts.append(mfr_pn)
        
        # Be polite to the API - small delay between requests
        time.sleep(0.5)
    
    # Save the populated BOM
    if output_file is None:
        base_name = os.path.splitext(bom_file_path)[0]
        output_file = f"{base_name}_populated.xlsx"
    
    print("\n" + "="*70)
    print("💾 Saving populated BOM...")
    try:
        workbook.save(output_file)
        print(f"✅ Saved to: {output_file}")
    except Exception as e:
        print(f"❌ Error saving file: {e}")
        return
    
    # Print summary
    print("\n" + "="*70)
    print("📊 SUMMARY")
    print("="*70)
    print(f"Total parts processed: {processed_count}")
    print(f"Successfully found: {success_count}")
    print(f"Not found: {len(failed_parts)}")
    
    if failed_parts:
        print(f"\n⚠️  Parts not found in DigiKey:")
        for part in failed_parts:
            print(f"   - {part}")
    
    print(f"\n✅ BOM population complete!")
    print(f"📁 Output file: {output_file}")
    print("="*70)


# ============================================================
# MAIN PROGRAM
# ============================================================

if __name__ == "__main__":
    
    # DigiKey API Credentials
    DIGIKEY_CLIENT_ID = "ENciL4CxjO0MAfsCMveTtNb2Ke96Xi5vSt5vanaouZIQbTq5"
    DIGIKEY_CLIENT_SECRET = "gGRJn44v69tnXdYUUEM27ELExG64yG4LoyQdVKegJlpDHTvDbXU87cevg6HvHIoo"
    
    # Input BOM file
    BOM_FILE = "Cable_BOM.xlsx"
    
    # Output file (set to None to auto-generate name)
    OUTPUT_FILE = "Cable_BOM_populated.xlsx"
    
    # Run the population
    populate_bom_from_digikey(
        bom_file_path=BOM_FILE,
        client_id=DIGIKEY_CLIENT_ID,
        client_secret=DIGIKEY_CLIENT_SECRET,
        output_file=OUTPUT_FILE
    )
