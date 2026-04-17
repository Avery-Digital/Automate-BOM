import customtkinter as ctk
import threading
import queue
import os
import openpyxl

from gui.frames.input_frame import InputFrame
from gui.frames.settings_frame import SettingsDialog
from gui.frames.qty_settings_frame import QtySettingsDialog
from gui.frames.progress_frame import ProgressFrame
from core.config import load_config
from core.bom_populator import BOMPopulator
from core.bom_builder import create_workbook
from core.importers import excel_importer, altium_importer, csv_importer


class BOMApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("BOM Auto-Populator")
        self.geometry("850x700")
        self.minsize(750, 600)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self._queue = queue.Queue()
        self._worker_thread = None
        self._populator = None
        self._last_output_file = None

        # Title
        title_frame = ctk.CTkFrame(self, fg_color="transparent")
        title_frame.grid(row=0, column=0, padx=20, pady=(15, 5))

        title = ctk.CTkLabel(
            title_frame, text="BOM Auto-Populator",
            font=("Arial", 22, "bold"))
        title.pack()

        subtitle = ctk.CTkLabel(
            title_frame, text="DigiKey Part Lookup & BOM Population",
            font=("Arial", 12), text_color="#AAAAAA")
        subtitle.pack(pady=(2, 0))

        # Input frame
        self.input_frame = InputFrame(
            self, on_settings_click=self._open_settings,
            on_qty_settings_click=self._open_qty_settings)
        self.input_frame.grid(row=1, column=0, padx=15, pady=8, sticky="ew")

        # Progress frame
        self.progress_frame = ProgressFrame(self)
        self.progress_frame.grid(row=2, column=0, padx=15, pady=8, sticky="nsew")

        # Button row
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.grid(row=3, column=0, padx=15, pady=(5, 15))

        self.run_btn = ctk.CTkButton(
            btn_frame, text="Run", width=160, height=40,
            font=("Arial", 14, "bold"), command=self._on_run)
        self.run_btn.pack(side="left", padx=10)

        self.cancel_btn = ctk.CTkButton(
            btn_frame, text="Cancel", width=160, height=40,
            font=("Arial", 14, "bold"), state="disabled",
            fg_color="#CC3333", hover_color="#AA2222",
            command=self._on_cancel)
        self.cancel_btn.pack(side="left", padx=10)

        # Export buttons frame (row below main buttons)
        export_frame = ctk.CTkFrame(self, fg_color="transparent")
        export_frame.grid(row=4, column=0, padx=15, pady=(0, 10))

        self.export_dk_btn = ctk.CTkButton(
            export_frame, text="Export DigiKey Cart", width=180, height=36,
            font=("Arial", 12, "bold"), state="disabled",
            fg_color="#2E7D32", hover_color="#388E3C",
            command=lambda: self._on_export_cart('DigiKey'))
        self.export_dk_btn.pack(side="left", padx=8)

        self.export_mouser_btn = ctk.CTkButton(
            export_frame, text="Export Mouser Cart", width=160, height=36,
            font=("Arial", 12, "bold"), state="disabled",
            fg_color="#2E7D32", hover_color="#388E3C",
            command=lambda: self._on_export_cart('Mouser'))
        self.export_mouser_btn.pack(side="left", padx=6)

        self.export_newark_btn = ctk.CTkButton(
            export_frame, text="Export Newark Cart", width=160, height=36,
            font=("Arial", 12, "bold"), state="disabled",
            fg_color="#2E7D32", hover_color="#388E3C",
            command=lambda: self._on_export_cart('Newark'))
        self.export_newark_btn.pack(side="left", padx=6)

        # Poll queue
        self._poll_queue()

        # Handle window close
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _open_settings(self):
        SettingsDialog(self)

    def _open_qty_settings(self):
        QtySettingsDialog(self)

    def _on_run(self):
        inputs = self.input_frame.get_inputs()

        # Validate
        if not inputs['bom_name']:
            self.progress_frame.append_log("ERROR: Please enter a BOM name.")
            return
        if not inputs['file_path']:
            self.progress_frame.append_log("ERROR: Please select an input file.")
            return

        config = load_config()
        if not config['digikey_client_id'] or not config['digikey_client_secret']:
            self.progress_frame.append_log("ERROR: DigiKey credentials not configured. Click Settings.")
            return

        # Reset UI
        self.progress_frame.reset()
        self.run_btn.configure(state="disabled")
        self.cancel_btn.configure(state="normal")

        # Start worker
        self._worker_thread = threading.Thread(
            target=self._run_worker, args=(inputs, config), daemon=True)
        self._worker_thread.start()

    def _run_worker(self, inputs, config):
        try:
            mode = inputs['mode']
            file_path = inputs['file_path']
            bom_name = inputs['bom_name']
            num_boards = inputs.get('num_boards', 1)

            distributor = inputs.get('distributor', 'DigiKey')

            populator = BOMPopulator(
                client_id=config['digikey_client_id'],
                client_secret=config['digikey_client_secret'],
                mouser_api_key=config.get('mouser_api_key', ''),
                newark_api_key=config.get('newark_api_key', ''),
                distributor=distributor,
                qty_settings=config.get('qty_settings'),
                log_callback=lambda msg: self._queue.put(('log', msg)),
                progress_callback=lambda c, t, pn, s: self._queue.put(('progress', c, t, pn, s)),
            )
            self._populator = populator

            workbook = None
            bom_file_path = None

            if mode == "Excel BOM":
                valid, err = excel_importer.validate_excel_bom(file_path)
                if not valid:
                    self._queue.put(('log', f"ERROR: {err}"))
                    self._queue.put(('done', None))
                    return
                bom_file_path = file_path

            elif mode == "Altium BOM":
                self._queue.put(('log', f"Importing Altium BOM: {file_path}"))
                parts = altium_importer.parse(file_path)
                self._queue.put(('log', f"Found {len(parts)} parts in Altium BOM"))
                if not parts:
                    self._queue.put(('log', "ERROR: No parts found in Altium BOM"))
                    self._queue.put(('done', None))
                    return
                workbook = create_workbook(bom_name, parts, num_boards=num_boards)

            elif mode == "Part Number List":
                self._queue.put(('log', f"Importing CSV: {file_path}"))
                parts = csv_importer.parse(file_path)
                self._queue.put(('log', f"Found {len(parts)} parts in CSV"))
                if not parts:
                    self._queue.put(('log', "ERROR: No parts found in CSV"))
                    self._queue.put(('done', None))
                    return
                workbook = create_workbook(bom_name, parts, num_boards=num_boards)

            # Build output filename from BOM name
            safe_name = "".join(c for c in bom_name if c not in '<>:"/\\|?*')
            output_file = f"{safe_name}_Populated.xlsx"

            stats = populator.populate_bom(
                bom_file_path=bom_file_path,
                workbook=workbook,
                output_file=output_file,
                bom_name=bom_name,
                num_boards=num_boards,
            )

            self._queue.put(('done', stats, output_file))

        except Exception as e:
            self._queue.put(('log', f"ERROR: {e}"))
            self._queue.put(('done', None, None))

    def _on_cancel(self):
        if self._populator:
            self._populator.cancel()
            self.progress_frame.append_log("Cancelling...")

    def _on_export_cart(self, target_distributor='DigiKey'):
        if not self._last_output_file:
            self.progress_frame.append_log("ERROR: No populated BOM to export from. Run the tool first.")
            return

        try:
            import csv
            from tkinter import filedialog
            wb = openpyxl.load_workbook(self._last_output_file)
            ws = wb.active

            # Find header row and columns
            header_row = None
            dist_pn_col = None
            distributor_col = None
            qty_to_buy_col = None
            available_col = None
            mfr_pn_col = None

            for row_idx in range(1, 11):
                for col_idx, cell in enumerate(ws[row_idx], start=1):
                    val = str(cell.value).strip() if cell.value else ""
                    if "Mfr" in val and "P/N" in val:
                        header_row = row_idx
                        break
                if header_row:
                    break

            if header_row:
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    val = str(cell.value).strip() if cell.value else ""
                    if "Dist" in val and "P/N" in val:
                        dist_pn_col = col_idx
                    elif "Distribut" in val and "P/N" not in val:
                        distributor_col = col_idx
                    elif "QTY TO BUY" in val.upper():
                        qty_to_buy_col = col_idx
                    elif "Available" in val:
                        available_col = col_idx
                    elif "Quantity Total" in val and not qty_to_buy_col:
                        qty_to_buy_col = col_idx
                    elif "Mfr" in val and "P/N" in val:
                        mfr_pn_col = col_idx

            if not dist_pn_col:
                self.progress_frame.append_log("ERROR: Could not find 'Dist. P/N' column.")
                return

            # Collect cart items
            cart_items = []
            skipped = 0
            for row_idx in range(header_row + 1, ws.max_row + 1):
                pn = ws.cell(row=row_idx, column=dist_pn_col).value
                if not pn or str(pn).strip() in ('', 'NOT FOUND', 'None'):
                    continue

                # Filter by distributor if column exists
                if distributor_col:
                    dist_val = str(ws.cell(row=row_idx, column=distributor_col).value or '').strip()
                    if dist_val and dist_val != target_distributor:
                        skipped += 1
                        continue

                # Skip 0 available
                if available_col:
                    avail = ws.cell(row=row_idx, column=available_col).value
                    try:
                        if avail is not None and int(avail) == 0:
                            continue
                    except (ValueError, TypeError):
                        pass

                # Get quantity
                qty = 1
                if qty_to_buy_col:
                    qty_val = ws.cell(row=row_idx, column=qty_to_buy_col).value
                    if qty_val:
                        try:
                            qty = int(qty_val)
                        except (ValueError, TypeError):
                            qty = 1
                if qty <= 0:
                    qty = 1

                cart_items.append((str(pn).strip(), qty))

            if not cart_items:
                self.progress_frame.append_log(f"No {target_distributor} items to export.")
                return

            # Save CSV
            suffix = f"_{target_distributor}_Cart.csv"
            default_name = os.path.splitext(self._last_output_file)[0] + suffix
            save_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                initialfile=os.path.basename(default_name),
            )
            if not save_path:
                return

            with open(save_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if target_distributor == 'Mouser':
                    writer.writerow(['Mouser Part Number', 'Quantity'])
                elif target_distributor == 'Newark':
                    writer.writerow(['Newark Part Number', 'Quantity'])
                else:
                    writer.writerow(['DigiKey Part Number', 'Quantity'])
                for pn, qty in cart_items:
                    writer.writerow([pn, qty])

            self.progress_frame.append_log(f"Exported {len(cart_items)} {target_distributor} items to: {save_path}")
            if skipped:
                self.progress_frame.append_log(f"  ({skipped} items from other distributor skipped)")
            upload_sites = {'DigiKey': 'DigiKey BOM Manager', 'Mouser': 'Mouser BOM Tool', 'Newark': 'Newark BOM Tool'}
            upload_site = upload_sites.get(target_distributor, target_distributor)
            self.progress_frame.append_log(f"Upload this CSV to {upload_site} to fill your cart.")

        except Exception as e:
            self.progress_frame.append_log(f"ERROR exporting cart: {e}")

    def _poll_queue(self):
        found = 0
        not_found = 0

        try:
            while True:
                item = self._queue.get_nowait()

                if item[0] == 'log':
                    self.progress_frame.append_log(item[1])

                elif item[0] == 'progress':
                    _, current, total, pn, status = item
                    self.progress_frame.update_progress(current, total, pn, status)

                elif item[0] == 'done':
                    stats = item[1]
                    output_file = item[2] if len(item) > 2 else None
                    if stats:
                        self.progress_frame.update_stats(
                            stats['success'], stats['failed'], stats['skipped'])
                        self.progress_frame.set_complete(stats)
                    if output_file:
                        self._last_output_file = output_file
                        self.export_dk_btn.configure(state="normal")
                        self.export_mouser_btn.configure(state="normal")
                        self.export_newark_btn.configure(state="normal")
                    self.run_btn.configure(state="normal")
                    self.cancel_btn.configure(state="disabled")
                    self._populator = None

        except queue.Empty:
            pass

        self.after(100, self._poll_queue)

    def _on_close(self):
        if self._populator:
            self._populator.cancel()
        self.destroy()
